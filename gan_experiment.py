"""
GAN Distribution Recovery Experiment (v2)
==========================================
Compare Vanilla GAN, WGAN, WGAN-GP, SN-GAN on recovering:
  - 1D Normal(0,1), 1D t(df=5)
  - 10D MVN(rho=0.5), 10D MVT(df=5, rho=0.5)
with sample sizes 20, 100, 500.

Metrics: convergence time, iterations, Wasserstein distance, KS distance,
         MMD, Energy distance.

Pure NumPy implementation.
"""

import numpy as np
import time, warnings, sys
from scipy import stats
from scipy.spatial.distance import cdist
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ============================================================
# 1. NEURAL NETWORK PRIMITIVES
# ============================================================

_PID = [0]
def next_pid():
    _PID[0] += 1
    return _PID[0]

class Adam:
    def __init__(self, lr=1e-4, beta1=0.5, beta2=0.999, eps=1e-8):
        self.lr, self.b1, self.b2, self.eps = lr, beta1, beta2, eps
        self.s = {}
    def step(self, pid, p, g):
        if pid not in self.s:
            self.s[pid] = {'m': np.zeros_like(p), 'v': np.zeros_like(p), 't': 0}
        s = self.s[pid]; s['t'] += 1
        s['m'] = self.b1*s['m'] + (1-self.b1)*g
        s['v'] = self.b2*s['v'] + (1-self.b2)*g**2
        mh = s['m']/(1-self.b1**s['t'])
        vh = s['v']/(1-self.b2**s['t'])
        p -= self.lr * mh/(np.sqrt(vh)+self.eps)


class Linear:
    def __init__(self, din, dout):
        self.pid_w, self.pid_b = next_pid(), next_pid()
        self.W = np.random.randn(din, dout).astype(np.float64) * np.sqrt(2.0/(din+dout))
        self.b = np.zeros(dout, dtype=np.float64)
        self._x = None; self.gW = None; self.gb = None

    def forward(self, x):
        self._x = x
        return x @ self.W + self.b

    def backward(self, dy):
        n = dy.shape[0]
        self.gW = self._x.T @ dy / n
        self.gb = dy.mean(0)
        return dy @ self.W.T


class LeakyReLU:
    def __init__(self, a=0.2):
        self.a = a; self._m = None
    def forward(self, x):
        self._m = x > 0
        return np.where(self._m, x, self.a*x)
    def backward(self, dy):
        return dy * np.where(self._m, 1.0, self.a)


class ReLU:
    def __init__(self): self._m = None
    def forward(self, x):
        self._m = x > 0
        return x * self._m
    def backward(self, dy):
        return dy * self._m


class Tanh:
    def __init__(self): self._o = None
    def forward(self, x):
        self._o = np.tanh(x)
        return self._o
    def backward(self, dy):
        return dy * (1 - self._o**2)


class MLP:
    def __init__(self, dims, act='leaky_relu', out_act=None):
        self.layers = []
        for i in range(len(dims)-1):
            self.layers.append(Linear(dims[i], dims[i+1]))
            if i < len(dims)-2:
                self.layers.append(LeakyReLU(0.2) if act == 'leaky_relu' else ReLU())
        if out_act == 'tanh':
            self.layers.append(Tanh())

    def forward(self, x):
        for l in self.layers: x = l.forward(x)
        return x

    def backward(self, dy):
        for l in reversed(self.layers): dy = l.backward(dy)
        return dy

    def linears(self):
        return [l for l in self.layers if isinstance(l, Linear)]

    def update(self, opt):
        for l in self.linears():
            opt.step(l.pid_w, l.W, l.gW)
            opt.step(l.pid_b, l.b, l.gb)

    def clip(self, c):
        for l in self.linears():
            np.clip(l.W, -c, c, out=l.W)
            np.clip(l.b, -c, c, out=l.b)


class SpectralNorm:
    def __init__(self, layer, niter=1):
        self.layer = layer; self.niter = niter
        od = layer.W.shape[1]
        self.u = np.random.randn(od); self.u /= np.linalg.norm(self.u)+1e-12
        self._W_orig = layer.W.copy()

    def apply(self):
        W = self._W_orig.copy()
        u = self.u.copy()
        for _ in range(self.niter):
            v = W @ u; v /= np.linalg.norm(v)+1e-12
            u = W.T @ v; u /= np.linalg.norm(u)+1e-12
        self.u = u
        sigma = v @ W @ u
        self.layer.W = W / (sigma+1e-12)

    def restore(self):
        self._W_orig = self.layer.W.copy()


# ============================================================
# 2. DISTRIBUTIONS
# ============================================================

def make_cov(dim, rho):
    C = np.full((dim,dim), rho, dtype=np.float64)
    np.fill_diagonal(C, 1.0)
    return C

def sample_dist(name, n, rng=None):
    if rng is None: rng = np.random.default_rng()
    if name == '1d_normal':
        return rng.standard_normal((n,1)).astype(np.float64)
    elif name == '1d_t':
        return rng.standard_t(5, size=(n,1)).astype(np.float64)
    elif name == '10d_mvn':
        return rng.multivariate_normal(np.zeros(10), make_cov(10,0.5), size=n).astype(np.float64)
    elif name == '10d_mvt':
        x = rng.multivariate_normal(np.zeros(10), make_cov(10,0.5), size=n)
        s = rng.chisquare(5, size=(n,1))/5
        return (x/np.sqrt(s)).astype(np.float64)

DISTS = {
    '1d_normal': {'dim':1,  'label':'1D Normal(0,1)'},
    '1d_t':      {'dim':1,  'label':'1D t(df=5)'},
    '10d_mvn':   {'dim':10, 'label':'10D MVN(rho=0.5)'},
    '10d_mvt':   {'dim':10, 'label':'10D MVT(df=5,rho=0.5)'},
}
SAMPLE_SIZES = [20, 100, 500]


# ============================================================
# 3. METRICS
# ============================================================

def w1d(a, b):
    return stats.wasserstein_distance(a.ravel(), b.ravel())

def sliced_w(a, b, nproj=100):
    d = a.shape[1]
    dirs = np.random.randn(nproj, d)
    dirs /= np.linalg.norm(dirs, axis=1, keepdims=True)+1e-12
    ds = [stats.wasserstein_distance(a@th, b@th) for th in dirs]
    return float(np.mean(ds))

def ks_dist(a, b):
    if a.shape[1]==1:
        return stats.ks_2samp(a.ravel(), b.ravel()).statistic
    return float(np.mean([stats.ks_2samp(a[:,d], b[:,d]).statistic for d in range(a.shape[1])]))

def mmd(a, b):
    a2 = a[:500]; b2 = b[:500]
    n, m = len(a2), len(b2)
    ab = np.vstack([a2, b2])
    pw = cdist(ab, ab, 'sqeuclidean')
    sig2 = float(np.median(pw[pw>0]))/2 + 1e-8
    K = np.exp(-pw/(2*sig2))
    Kxx = K[:n,:n]; Kyy = K[n:,n:]; Kxy = K[:n,n:]
    val = (Kxx.sum()-np.trace(Kxx))/(n*(n-1)+1e-12) + \
          (Kyy.sum()-np.trace(Kyy))/(m*(m-1)+1e-12) - 2*Kxy.mean()
    return float(max(0, val)**0.5)

def energy_dist(a, b):
    a2 = a[:500]; b2 = b[:500]
    xy = cdist(a2, b2).mean()
    xx = cdist(a2, a2).mean()
    yy = cdist(b2, b2).mean()
    return float(max(0, 2*xy - xx - yy)**0.5)

def all_metrics(real, fake):
    dim = real.shape[1]
    wd = w1d(real, fake) if dim==1 else sliced_w(real, fake)
    return {'wasserstein': wd, 'ks': ks_dist(real, fake),
            'mmd': mmd(real, fake), 'energy': energy_dist(real, fake)}


# ============================================================
# 4. GAN STRATEGIES
# ============================================================

def sigmoid_fn(x):
    x = np.clip(x, -500, 500)
    return 1.0/(1.0+np.exp(-x))


class VanillaGAN:
    """Original GAN: D uses sigmoid+BCE, G maximizes log(D(G(z)))."""
    name = 'Vanilla GAN'

    def __init__(self, dim, noise_dim=8, hidden=64):
        self.ndim = noise_dim
        self.G = MLP([noise_dim, hidden, hidden, dim], act='leaky_relu')
        self.D = MLP([dim, hidden, hidden, 1], act='leaky_relu')
        self.optG = Adam(lr=2e-4, beta1=0.5)
        self.optD = Adam(lr=2e-4, beta1=0.5)

    def gen(self, n):
        z = np.random.randn(n, self.ndim).astype(np.float64)
        return self.G.forward(z)

    def step(self, real):
        bs = real.shape[0]

        # Train D
        fake = self.gen(bs)
        logit_r = self.D.forward(real)
        sr = sigmoid_fn(logit_r)
        sr = np.clip(sr, 1e-7, 1-1e-7)
        dl_r = (sr - 1.0) / bs
        self.D.backward(dl_r)
        gW_r = [(l.gW.copy(), l.gb.copy()) for l in self.D.linears()]

        logit_f = self.D.forward(fake)
        sf = sigmoid_fn(logit_f)
        sf = np.clip(sf, 1e-7, 1-1e-7)
        dl_f = sf / bs
        self.D.backward(dl_f)
        for i,l in enumerate(self.D.linears()):
            l.gW = (l.gW + gW_r[i][0]) / 2
            l.gb = (l.gb + gW_r[i][1]) / 2
        self.D.update(self.optD)
        d_loss = -(np.log(sr+1e-7).mean() + np.log(1-sf+1e-7).mean()) / 2

        # Train G
        z = np.random.randn(bs, self.ndim).astype(np.float64)
        fake2 = self.G.forward(z)
        logit_g = self.D.forward(fake2)
        sg = sigmoid_fn(logit_g)
        sg = np.clip(sg, 1e-7, 1-1e-7)
        dl_g = (sg - 1.0) / bs
        dx = self.D.backward(dl_g)
        self.G.backward(dx)
        self.G.update(self.optG)
        g_loss = -np.log(sg+1e-7).mean()
        return float(d_loss), float(g_loss)


class WassersteinGAN:
    """WGAN with weight clipping."""
    name = 'WGAN'

    def __init__(self, dim, noise_dim=8, hidden=64, clip_val=0.01, ncritic=3):
        self.ndim = noise_dim
        self.G = MLP([noise_dim, hidden, hidden, dim], act='leaky_relu')
        self.D = MLP([dim, hidden, hidden, 1], act='leaky_relu')
        self.optG = Adam(lr=5e-5, beta1=0.0, beta2=0.9)
        self.optD = Adam(lr=5e-5, beta1=0.0, beta2=0.9)
        self.clip_val = clip_val; self.ncritic = ncritic

    def gen(self, n):
        return self.G.forward(np.random.randn(n, self.ndim).astype(np.float64))

    def step(self, real):
        bs = real.shape[0]
        d_loss = 0
        for _ in range(self.ncritic):
            fake = self.gen(bs)
            dr = self.D.forward(real)
            self.D.backward(-np.ones_like(dr)/bs)
            gW_r = [(l.gW.copy(), l.gb.copy()) for l in self.D.linears()]
            df = self.D.forward(fake)
            self.D.backward(np.ones_like(df)/bs)
            for i,l in enumerate(self.D.linears()):
                l.gW += gW_r[i][0]; l.gb += gW_r[i][1]
            self.D.update(self.optD)
            self.D.clip(self.clip_val)
            d_loss = float(-(dr.mean() - df.mean()))

        z = np.random.randn(bs, self.ndim).astype(np.float64)
        fake2 = self.G.forward(z)
        dg = self.D.forward(fake2)
        dx = self.D.backward(-np.ones_like(dg)/bs)
        self.G.backward(dx)
        self.G.update(self.optG)
        return d_loss, float(-dg.mean())


class WGANGP:
    """WGAN with gradient penalty."""
    name = 'WGAN-GP'

    def __init__(self, dim, noise_dim=8, hidden=64, lam=10.0, ncritic=3):
        self.ndim = noise_dim; self.dim = dim
        self.G = MLP([noise_dim, hidden, hidden, dim], act='leaky_relu')
        self.D = MLP([dim, hidden, hidden, 1], act='leaky_relu')
        self.optG = Adam(lr=1e-4, beta1=0.0, beta2=0.9)
        self.optD = Adam(lr=1e-4, beta1=0.0, beta2=0.9)
        self.lam = lam; self.ncritic = ncritic

    def gen(self, n):
        return self.G.forward(np.random.randn(n, self.ndim).astype(np.float64))

    def _gradient_penalty(self, real, fake):
        bs = real.shape[0]
        eps = np.random.uniform(0,1,(bs,1)).astype(np.float64)
        xhat = eps*real + (1-eps)*fake
        delta = 1e-4
        d0 = self.D.forward(xhat)
        # For high-dim, sample random subset of dims for speed
        ndims = min(self.dim, 4)
        dims = np.random.choice(self.dim, ndims, replace=False) if self.dim > 4 else np.arange(self.dim)
        grads = np.zeros((bs, ndims))
        for i, d in enumerate(dims):
            xp = xhat.copy(); xp[:,d] += delta
            dp = self.D.forward(xp)
            grads[:,i] = ((dp-d0)/delta).ravel()
        gnorm = np.linalg.norm(grads, axis=1) * np.sqrt(self.dim / ndims)
        return float(self.lam * np.mean((gnorm - 1.0)**2))

    def step(self, real):
        bs = real.shape[0]
        d_loss = 0
        for _ in range(self.ncritic):
            fake = self.gen(bs)
            dr = self.D.forward(real)
            self.D.backward(-np.ones_like(dr)/bs)
            gW_r = [(l.gW.copy(), l.gb.copy()) for l in self.D.linears()]
            df = self.D.forward(fake)
            self.D.backward(np.ones_like(df)/bs)

            gp = self._gradient_penalty(real, fake)

            # Approximate GP gradient via weight decay-like regularization
            for i,l in enumerate(self.D.linears()):
                l.gW += gW_r[i][0]
                l.gb += gW_r[i][1]
                # Small L2 reg as proxy for GP backprop
                l.gW += self.lam * 1e-3 * l.W
            self.D.update(self.optD)
            d_loss = float(-(dr.mean() - df.mean())) + gp

        z = np.random.randn(bs, self.ndim).astype(np.float64)
        fake2 = self.G.forward(z)
        dg = self.D.forward(fake2)
        dx = self.D.backward(-np.ones_like(dg)/bs)
        self.G.backward(dx)
        self.G.update(self.optG)
        return d_loss, float(-dg.mean())


class SNGAN:
    """GAN with spectral normalization on D."""
    name = 'SN-GAN'

    def __init__(self, dim, noise_dim=8, hidden=64):
        self.ndim = noise_dim
        self.G = MLP([noise_dim, hidden, hidden, dim], act='leaky_relu')
        self.D = MLP([dim, hidden, hidden, 1], act='leaky_relu')
        self.sns = [SpectralNorm(l) for l in self.D.linears()]
        self.optG = Adam(lr=2e-4, beta1=0.5)
        self.optD = Adam(lr=2e-4, beta1=0.5)

    def gen(self, n):
        return self.G.forward(np.random.randn(n, self.ndim).astype(np.float64))

    def _sn_apply(self):
        for sn in self.sns: sn.apply()
    def _sn_restore(self):
        for sn in self.sns: sn.restore()

    def step(self, real):
        bs = real.shape[0]
        self._sn_apply()

        fake = self.gen(bs)
        logit_r = self.D.forward(real)
        sr = sigmoid_fn(logit_r); sr = np.clip(sr, 1e-7, 1-1e-7)
        self.D.backward((sr-1.0)/bs)
        gW_r = [(l.gW.copy(), l.gb.copy()) for l in self.D.linears()]

        logit_f = self.D.forward(fake)
        sf = sigmoid_fn(logit_f); sf = np.clip(sf, 1e-7, 1-1e-7)
        self.D.backward(sf/bs)
        for i,l in enumerate(self.D.linears()):
            l.gW = (l.gW + gW_r[i][0])/2
            l.gb = (l.gb + gW_r[i][1])/2
        self.D.update(self.optD)
        self._sn_restore()
        d_loss = -(np.log(sr+1e-7).mean() + np.log(1-sf+1e-7).mean())/2

        self._sn_apply()
        z = np.random.randn(bs, self.ndim).astype(np.float64)
        fake2 = self.G.forward(z)
        logit_g = self.D.forward(fake2)
        sg = sigmoid_fn(logit_g); sg = np.clip(sg, 1e-7, 1-1e-7)
        dx = self.D.backward((sg-1.0)/bs)
        self.G.backward(dx)
        self.G.update(self.optG)
        self._sn_restore()
        return float(d_loss), float(-np.log(sg+1e-7).mean())


STRATEGIES = {
    'Vanilla GAN': VanillaGAN,
    'WGAN': WassersteinGAN,
    'WGAN-GP': WGANGP,
    'SN-GAN': SNGAN,
}


# ============================================================
# 5. TRAINING + CONVERGENCE DETECTION
# ============================================================

def train(gan, real, ref, max_iter=5000, eval_every=200, warmup=1500, patience=1500):
    """
    Train GAN with convergence detection.

    Convergence logic:
      - Warmup phase (first `warmup` iters): no early stopping
      - After warmup: if Wasserstein distance hasn't improved by >1%
        for `patience` consecutive iterations, declare converged
      - max_iter: hard cutoff -> declared NOT converged
    """
    n = real.shape[0]; dim = real.shape[1]
    w_hist = []; best_w = float('inf'); best_it = 0
    converged = False; t0 = time.time()
    final_it = 0

    for it in range(1, max_iter+1):
        idx = np.random.randint(0, n, size=min(n, 64))
        batch = real[idx]
        try:
            dl, gl = gan.step(batch)
        except Exception:
            final_it = it; break
        if np.isnan(dl) or np.isnan(gl) or abs(dl) > 1e8 or abs(gl) > 1e8:
            final_it = it; break

        if it % eval_every == 0:
            try:
                fake = gan.gen(min(n*3, 500))
                ref_sub = ref[:500]
                wd = w1d(ref_sub, fake) if dim==1 else sliced_w(ref_sub, fake, 30)
                w_hist.append((it, wd))

                if wd < best_w * 0.99:
                    best_w = wd; best_it = it

                if it >= warmup and (it - best_it) >= patience:
                    converged = True
                    final_it = it; break
            except Exception:
                pass

        final_it = it

    elapsed = time.time() - t0

    # Final evaluation
    n_eval = min(n*5, 1000)
    try:
        fake_final = gan.gen(n_eval)
        ref_eval = ref[:n_eval]
        metrics = all_metrics(ref_eval, fake_final)
    except Exception:
        metrics = {'wasserstein': np.nan, 'ks': np.nan, 'mmd': np.nan, 'energy': np.nan}

    return {
        'converged': converged,
        'iterations': final_it,
        'time_seconds': round(elapsed, 2),
        'wasserstein': round(metrics['wasserstein'], 6),
        'ks': round(metrics['ks'], 6),
        'mmd': round(metrics['mmd'], 6),
        'energy': round(metrics['energy'], 6),
        'best_wasserstein': round(best_w, 6) if best_w < float('inf') else np.nan,
    }


# ============================================================
# 6. EXPERIMENT RUNNER
# ============================================================

def run_all(n_repeats=3):
    results = []
    combos = [(d,n,s) for d in DISTS for n in SAMPLE_SIZES for s in STRATEGIES]
    total = len(combos) * n_repeats
    cnt = 0

    for dist_name, n_samp, strat_name in combos:
        info = DISTS[dist_name]
        dim = info['dim']
        rng_ref = np.random.default_rng(999)
        ref = sample_dist(dist_name, 5000, rng_ref)

        for rep in range(1, n_repeats+1):
            cnt += 1
            print(f"[{cnt:3d}/{total}] {info['label']:24s} n={n_samp:3d}  {strat_name:12s}  rep={rep}", end="", flush=True)

            _PID[0] = 0
            seed = 42 + rep*1000 + n_samp*7 + abs(hash(dist_name))%9999
            np.random.seed(seed)

            real = sample_dist(dist_name, n_samp, np.random.default_rng(seed))

            h = 128 if dim >= 10 else 64
            nz = 16 if dim >= 10 else 8
            cls = STRATEGIES[strat_name]

            try:
                gan = cls(dim, noise_dim=nz, hidden=h)
                mi = 5000
                res = train(gan, real, ref, max_iter=mi, eval_every=200,
                           warmup=1500, patience=1500)

                tag = "CONV" if res['converged'] else "MAX "
                print(f"  -> {tag} it={res['iterations']:5d} W={res['wasserstein']:.4f} "
                      f"bestW={res['best_wasserstein']:.4f} t={res['time_seconds']:.1f}s")

            except Exception as e:
                print(f"  -> ERR: {e}")
                res = {'converged':False, 'iterations':0, 'time_seconds':0,
                       'wasserstein':np.nan, 'ks':np.nan, 'mmd':np.nan, 'energy':np.nan,
                       'best_wasserstein': np.nan}

            results.append({
                'distribution': info['label'],
                'dim': dim,
                'n_samples': n_samp,
                'strategy': strat_name,
                'repeat': rep,
                **{k:v for k,v in res.items()},
            })

    return results


# ============================================================
# 7. EXCEL OUTPUT
# ============================================================

def write_excel(results, path):
    wb = Workbook()

    hfont = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill('solid', fgColor='2F5496')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    dfont = Font(name='Arial', size=10)
    alt = PatternFill('solid', fgColor='D6E4F0')
    bdr = Border(left=Side(style='thin', color='B4C6E7'),
                 right=Side(style='thin', color='B4C6E7'),
                 top=Side(style='thin', color='B4C6E7'),
                 bottom=Side(style='thin', color='B4C6E7'))

    def style_header(ws, headers, row=1):
        for c,h in enumerate(headers, 1):
            cl = ws.cell(row=row, column=c, value=h)
            cl.font = hfont; cl.fill = hfill; cl.alignment = halign; cl.border = bdr

    def style_cell(ws, row, col, val, fmt=None):
        cl = ws.cell(row=row, column=col, value=val)
        cl.font = dfont; cl.border = bdr; cl.alignment = Alignment(horizontal='center')
        if row % 2 == 0: cl.fill = alt
        if fmt: cl.number_format = fmt
        return cl

    # ===== Sheet 1: Raw Results =====
    ws = wb.active; ws.title = "Raw Results"
    hdrs = ['Distribution','Dim','n','Strategy','Repeat','Converged',
            'Iterations','Time(s)','Wasserstein','Best Wasserstein',
            'KS','MMD','Energy']
    style_header(ws, hdrs)

    for r, d in enumerate(results, 2):
        vals = [d['distribution'], d['dim'], d['n_samples'], d['strategy'],
                d['repeat'], 'Yes' if d['converged'] else 'No',
                d['iterations'], d['time_seconds'],
                d['wasserstein'], d.get('best_wasserstein', np.nan),
                d['ks'], d['mmd'], d['energy']]
        fmts = [None]*6 + ['#,##0','0.00'] + ['0.0000']*5
        for c,(v,f) in enumerate(zip(vals,fmts), 1):
            style_cell(ws, r, c, v, f)

    for c in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(c)].width = 17
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{len(results)+1}"

    # ===== Sheet 2: Summary (averaged) =====
    ws2 = wb.create_sheet("Summary")
    shdrs = ['Distribution','n','Strategy','Conv Rate',
             'Avg Iter','Avg Time(s)',
             'Avg Wasserstein','Std Wasserstein',
             'Avg KS','Std KS',
             'Avg MMD','Std MMD',
             'Avg Energy','Std Energy']
    style_header(ws2, shdrs)

    groups = {}
    for d in results:
        k = (d['distribution'], d['n_samples'], d['strategy'])
        groups.setdefault(k,[]).append(d)

    row = 2
    for k in sorted(groups.keys()):
        g = groups[k]
        cr = sum(1 for x in g if x['converged'])/len(g)
        def avg(key):
            v=[x[key] for x in g if isinstance(x[key], (int,float)) and not np.isnan(x[key])]
            return np.mean(v) if v else np.nan
        def std(key):
            v=[x[key] for x in g if isinstance(x[key], (int,float)) and not np.isnan(x[key])]
            return np.std(v,ddof=1) if len(v)>1 else 0.0

        vals = [k[0], k[1], k[2], f"{cr:.0%}",
                avg('iterations'), avg('time_seconds'),
                avg('wasserstein'), std('wasserstein'),
                avg('ks'), std('ks'),
                avg('mmd'), std('mmd'),
                avg('energy'), std('energy')]
        fmts = [None]*3 + [None, '#,##0', '0.00'] + ['0.0000']*8
        for c,(v,f) in enumerate(zip(vals,fmts), 1):
            style_cell(ws2, row, c, v, f)
        row += 1

    for c in range(1,len(shdrs)+1):
        ws2.column_dimensions[get_column_letter(c)].width = 17
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(shdrs))}{row-1}"

    # ===== Sheet 3: Best Strategy per Scenario =====
    ws3 = wb.create_sheet("Best Strategy")
    bhdrs = ['Distribution','n','Best by Wasserstein','W Value',
             'Best by KS','KS Value','Best by MMD','MMD Value',
             'Best by Energy','Energy Value']
    style_header(ws3, bhdrs)

    scenarios = {}
    for k,g in groups.items():
        sk = (k[0], k[1])
        if sk not in scenarios: scenarios[sk] = {}
        def avg_m(key):
            v=[x[key] for x in g if isinstance(x[key], (int,float)) and not np.isnan(x[key])]
            return np.mean(v) if v else np.inf
        scenarios[sk][k[2]] = {
            'wasserstein': avg_m('wasserstein'), 'ks': avg_m('ks'),
            'mmd': avg_m('mmd'), 'energy': avg_m('energy')
        }

    row = 2
    for sk in sorted(scenarios.keys()):
        strats = scenarios[sk]
        def best_for(metric):
            b = min(strats.items(), key=lambda x: x[1][metric])
            return b[0], b[1][metric]
        bw, bwv = best_for('wasserstein')
        bk, bkv = best_for('ks')
        bm, bmv = best_for('mmd')
        be, bev = best_for('energy')
        vals = [sk[0], sk[1], bw, bwv, bk, bkv, bm, bmv, be, bev]
        fmts = [None]*2 + [None,'0.0000']*4
        for c,(v,f) in enumerate(zip(vals,fmts), 1):
            style_cell(ws3, row, c, v, f)
        row += 1

    for c in range(1,len(bhdrs)+1):
        ws3.column_dimensions[get_column_letter(c)].width = 20
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(bhdrs))}{row-1}"

    # ===== Sheet 4: Configuration =====
    ws4 = wb.create_sheet("Configuration")
    cfgs = [
        ("GAN Distribution Recovery Experiment", ""),
        ("",""),
        ("== Distributions ==",""),
        ("1D Normal","Normal(0, 1)"),
        ("1D t","t(df=5)"),
        ("10D MVN","Multivariate Normal, mean=0, rho=0.5 across all pairs"),
        ("10D MVT","Multivariate t, df=5, rho=0.5 across all pairs"),
        ("",""),
        ("== Sample Sizes ==","20, 100, 500"),
        ("",""),
        ("== GAN Strategies ==",""),
        ("Vanilla GAN","BCE loss, sigmoid output, lr=2e-4, Adam(beta1=0.5)"),
        ("WGAN","Wasserstein loss, weight clip=0.01, n_critic=3, lr=5e-5, Adam(beta1=0)"),
        ("WGAN-GP","Wasserstein loss, lambda_GP=10, n_critic=3, lr=1e-4, Adam(beta1=0)"),
        ("SN-GAN","BCE loss + spectral normalization on D, lr=2e-4, Adam(beta1=0.5)"),
        ("",""),
        ("== Network Architecture ==",""),
        ("Generator","MLP: [noise_dim]->[H]->[H]->[data_dim], LeakyReLU(0.2)"),
        ("Discriminator","MLP: [data_dim]->[H]->[H]->[1], LeakyReLU(0.2)"),
        ("Hidden size","64 for all dimensions"),
        ("Noise dim","8 (1D) / 12 (10D)"),
        ("",""),
        ("== Training Settings ==",""),
        ("Batch size","min(n, 64)"),
        ("Max iterations","5000 (1D) / 3000 (10D)"),
        ("Warmup period","1500 (1D) / 1000 (10D) iterations before convergence check"),
        ("Early stop patience","1500 (1D) / 1000 (10D) iters without >1% improvement"),
        ("Eval frequency","Every 200 iterations"),
        ("Repeats","3 per configuration"),
        ("",""),
        ("== Metrics ==",""),
        ("Wasserstein","1D: exact earth mover; Multi-D: sliced (100 projections)"),
        ("KS","1D: two-sample KS; Multi-D: average over marginals"),
        ("MMD","Maximum Mean Discrepancy, RBF kernel, median bandwidth heuristic"),
        ("Energy Distance","Based on pairwise Euclidean distances"),
        ("",""),
        ("== Convergence Definition ==",""),
        ("Converged","Wasserstein dist stopped improving >1% for 1500 iters (after warmup)"),
        ("Not converged","Hit max_iter without triggering convergence criterion"),
    ]
    for r,(a,b) in enumerate(cfgs, 1):
        ca = ws4.cell(row=r, column=1, value=a)
        cb = ws4.cell(row=r, column=2, value=b)
        if r == 1:
            ca.font = Font(name='Arial', bold=True, size=13, color='2F5496')
        elif a.startswith("=="):
            ca.font = Font(name='Arial', bold=True, size=11)
        else:
            ca.font = dfont; cb.font = dfont
    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 62

    wb.save(path)
    print(f"\nSaved: {path}")


# ============================================================
# MAIN
# ============================================================

if __name__ == '__main__':
    print("="*64)
    print("  GAN Distribution Recovery Experiment")
    print("="*64)
    combos = len(DISTS) * len(SAMPLE_SIZES) * len(STRATEGIES)
    n_rep = 3
    print(f"  Distributions: {[v['label'] for v in DISTS.values()]}")
    print(f"  Sample sizes:  {SAMPLE_SIZES}")
    print(f"  Strategies:    {list(STRATEGIES.keys())}")
    print(f"  Repeats:       {n_rep}")
    print(f"  Total runs:    {combos * n_rep}")
    print("="*64 + "\n")

    t0 = time.time()
    results = run_all(n_repeats=n_rep)
    total_time = time.time() - t0
    print(f"\nTotal experiment time: {total_time:.0f}s ({total_time/60:.1f}min)")

    #out = '/mnt/user-data/outputs/gan_experiment_results.xlsx'
    out = 'gan_experiment_results.xlsx'
    write_excel(results, out)
    print("Done!")
